from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime, date
wb = load_workbook('/Users/priyam/Coding Stuff/shoe_store.xlsx')
ws = wb['Sheet1']

class Shoe:
    def __init__(self, item_number, brand, model, colorway, size, price):
        self.item_num = item_number
        self.brand = brand
        self.model = model
        self.cw = colorway
        self.size = size
        self.price = price
    
    def __repr__(self) -> str:
        return f"(Item #: {self.item_num}) {self.brand} {self.model} {self.cw} (Size {self.size})"

class Inventory:
    def __init__(self):
        self.list_of_shoes = []  #[(shoe, qty), (shoe, qty)]
    
    def add(self, shoe, qty):
        if shoe.item_num in [shoe.item_num for shoe, qty in self.list_of_shoes]:
           for index, (existing_shoe, existing_qty) in enumerate(self.list_of_shoes):
               if shoe.item_num == existing_shoe.item_num:
                   self.list_of_shoes[index] = (shoe, self.list_of_shoes[index][1]+ qty)         
        else:
            self.list_of_shoes.append((shoe, qty))

    def show(self):
        for x in (self.list_of_shoes):
            shoe = x[0]
            qty = x[1]
            print(f"{qty} -> {shoe}")
    
class Actions:
    def add_new_shoe_to_inventory(self):
        item_num = input('Item #: ')
        brand = input('Brand: ')
        model = input('Model: ')
        cw = input('Colorway: ')
        sz = input('Size: ')
        price = int(input('Price: '))
        pairs = int(input('How many pairs: '))
        new_shoe = Shoe(item_num, brand, model, cw, sz, price)
        inventory.add(new_shoe, pairs)
        print('the inventory has been updated')
    
    def select_shoe_from_inventory(self):
        item_num = input('Which item #: ')
        for x in inventory.list_of_shoes:
            shoe = x[0]
            if shoe.item_num == item_num:
                selected_shoe = shoe
                print(selected_shoe)
        
    def remove(self):
        item_num = input('Which item #: ')
        for x in inventory.list_of_shoes:
            shoe = x[0]
            if shoe.item_num == item_num:
                selected_shoe = shoe
                verify = input(f"Is {selected_shoe} correct?")
                if verify.lower() == 'yes':
                    inventory.list_of_shoes.remove(x)
                else: print('okay')
        print('The shoe has been removed')
    
    def update_shoe(self):
        item_num = input('Which item #: ')
        for x in inventory.list_of_shoes:
            shoe = x[0]
            quantity = x[1]
            if shoe.item_num == item_num:
                selected_shoe = shoe
        update = input("\nWhat would you like to update?\n1.Item Number\n2.Brand\n3.Model\n4.Colorway\n5.Size\n6.Price\n7.Quantity\n")
        if update == '1':
            number = input('new item #: ')
            selected_shoe.item_num = number
            print(selected_shoe)
        if update == '2':
            brand = input('new brand name: ')
            selected_shoe.brand = brand
            print(selected_shoe)
        if update == '3':
            model = input('new model name: ')
            selected_shoe.model = model
            print(selected_shoe)
        if update == '4':
            cw = input('colorway: ')
            selected_shoe.cw = cw
            print(selected_shoe)
        if update == '5':
            size = input('size: ')
            selected_shoe.size = size
            print(selected_shoe)
        if update == '6':
            price = int(input('price: '))
            selected_shoe.price = price
            print(f"{selected_shoe}, New Price = {selected_shoe.price}")
        if update == '7':
            quantity = int(input('Quantity: '))
            inventory.add(selected_shoe, quantity)
    
    def sell(self):
        item_num = input('Which item #: ')
        for x in inventory.list_of_shoes:
            shoe = x[0]
            if shoe.item_num == item_num:
                selected_shoe = shoe
                total = int(input(f"How many pairs of {selected_shoe}? "))
                if shoe.item_num in [shoe.item_num for shoe, quantity in inventory.list_of_shoes]:
                    for index, (existing_shoe, qty) in enumerate(inventory.list_of_shoes):
                        if shoe.item_num == existing_shoe.item_num:
                            inventory.list_of_shoes[index] = (shoe, inventory.list_of_shoes[index][1]- total)
                print(f"The total sale price is: {selected_shoe.price * total}")
    
    def refund(self):
        item_num = input('Which item #: ')
        for x in inventory.list_of_shoes:
            shoe = x[0]
            if shoe.item_num == item_num:
                selected_shoe = shoe
                total = int(input(f"How many pairs of {selected_shoe}? "))
                if shoe.item_num in [shoe.item_num for shoe, quantity in inventory.list_of_shoes]:
                    for index, (existing_shoe, qty) in enumerate(inventory.list_of_shoes):
                        if shoe.item_num == existing_shoe.item_num:
                            inventory.list_of_shoes[index] = (shoe, inventory.list_of_shoes[index][1] + total)
                print(f"Total refund cost: {selected_shoe.price * total}")

inventory = Inventory()
action = Actions()

#SHOES FROM EXCEL BEING ADDED TO LIST:
for x in ws.iter_rows(min_row=2,
                        values_only=True):
    shoe = Shoe(item_number=str(x[0]),
                brand=str(x[1]),
                model=str(x[2]),
                colorway=str(x[3]),
                size=str(x[4]),
                price=int(x[5])
                )
    qty = int(x[6])
    inventory.add(shoe, qty)
inventory.show()

while True:
    start_up = input('\nP11Kicks\nWould you like to:\n1.Update inventory\n2.Sell shoes from inventory\n3.Give a refund\n4.View inventory\n5.Exit\n')
    if start_up == '1':
        inventory.show()
        update_choice = input('\nWould you like to:\n1.Add a new shoe to inventory\n2.Remove a shoe from inventory\n3.Update a current shoe in inventory\n')
        if update_choice == '1':
            action.add_new_shoe_to_inventory()
        elif update_choice == '2':
            shoe_to_remove = print('Which shoe would you like to remove?\n')
            inventory.show()
            action.remove()
        elif update_choice == '3':
            shoe_to_update = print('Which shoe would you like to update?\n')
            inventory.show()
            action.update_shoe()
        else: print('Pick a valid option')

    elif start_up == '2':
        inventory.show()
        print('Select the shoe you want to sell:\n')
        action.sell()

    elif start_up == '3':
        inventory.show()
        print('Select the shoe you want to refund:\n')
        action.refund()

    elif start_up == '4':
        print('\nP11 Kicks Inventory:\nQuantity -> Brand, Model, Colorway, Size')
        inventory.show()

    elif start_up == '5':
        Wb = Workbook()
        Ws = Wb.active
        headers = ['Item #','Brand','Model','Colorway','Size','Price','Quantity']
        Ws.append(headers)
        for x in inventory.list_of_shoes:
            shoe = x[0]
            quantity = x[1]
            data = [shoe.item_num, shoe.brand, shoe.model, shoe.cw, shoe.size, shoe.price, quantity]
            Ws.append(data)
            file = Wb.save('/Users/priyam/Coding Stuff/shoe_store'+datetime.now().strftime('%Y%m%d')+'.xlsx')
        break

    else: print('Pick a valid number for a valid option')


wb.remove(ws)
updated_sheet = wb.create_sheet('Sheet1')
headers = ['Item #','Brand','Model','Colorway','Size','Price','Quantity']
updated_sheet.append(headers)
for x in inventory.list_of_shoes:
    shoe = x[0]
    quantity = x[1]
    data = [shoe.item_num, shoe.brand, shoe.model, shoe.cw, shoe.size, shoe.price, quantity]
    updated_sheet.append(data)
wb.save(filename='shoe_store.xlsx')